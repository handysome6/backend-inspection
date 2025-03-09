import cloudComPy as cc
cc.initCC()

import cv2
import json
from icecream import ic
from pathlib import Path
from loguru import logger
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.drawing.image import Image

from .utils import timer
from .preprocess import preprocess
from .dxf_analyze import dxf_analyze
from .extract_slices import extract_slices
from .main_fit_bottom_line import extract_hole_line
from .main_cad_align import cad_align, draw_text_on_img
from .main_rotateToGetMinBB import rotate_and_get_bb_dim


ROOT_SAVE_PATH = r"D:\Inspection_data"
EXCEL_TEMPLATE = "./Data/results.xlsx"     # this path is excel template


@timer
def all_measurement(cloud_path, cad_model_path):
    # cloud_path = r"C:\workspace\Test\Project_Hebei\ply\20241211001.ply"
    # cad_model_path = r"C:\workspace\Test\Project_Hebei\dxf\20241211001.dxf"
    logger.info(f"开始处理点云: {cloud_path}, CAD模型: {cad_model_path}")

    assert Path(cloud_path).exists(), "点云文件不存在"
    assert Path(cad_model_path).exists(), "CAD模型文件不存在"

    # target save path
    save_path = Path(ROOT_SAVE_PATH) / Path(cloud_path).parent.stem
    save_path.mkdir(parents=True, exist_ok=True)

    # =================加载点云====================
    cloud = cc.loadPointCloud(cloud_path)
    if cloud is None:
        logger.error("None pc found")
        exit(1)
    logger.success(f"点云加载成功, 路径: {cloud_path}")



    # =================预处理====================
    new_cloud, dia_diff, k_diff = preprocess(cloud)
    logger.success(f"点云预处理成功")
    background_color = (255, 255, 255)



    # =================计算墙体尺寸====================
    dia1_cloud = new_cloud.cloneThis()
    dia2_cloud = new_cloud.cloneThis()
    xs_argmax_x1, degree_argmax_x1 = rotate_and_get_bb_dim(dia1_cloud, -85, -30, 1)
    xs_argmax_x2, degree_argmax_x2 = rotate_and_get_bb_dim(dia2_cloud, 30, 85, 1)
    dimensions,_ = dxf_analyze(cad_model_path)
    # ic(dimensions)

    left_height, error_left_height, right_height, error_right_height,_,_,_,_ = extract_slices(new_cloud, dimensions, processRepeatX=True, processRepeatY=False)
    _,_,_,_,top_width, error_top_width, bottom_width, error_bottom_width =  extract_slices(new_cloud, dimensions, processRepeatX=False, processRepeatY=True)

    wall_sizes  = {
        'left_height': left_height,
        'right_height': right_height,
        'top_width': top_width,
        'bottom_width': bottom_width,
        'dia_diff': dia_diff
        }
    ic(wall_sizes)

    errors = {
        'window_width': None,
        'window_height': None,
        'window_dia_diff': None,
        'door_width': None,
        'door_height': None,
        'door_dia_diff': None,
        'wall_left_height': error_left_height,
        'wall_right_height': error_right_height,
        'wall_top_width': error_top_width,
        'wall_bottom_width': error_bottom_width,
        'wall_dia_diff':None
    }
    window_sizes = {
        'window_height': None,
        'window_width': None,
        'window_dia_diff': None
    }
    door_sizes = {
        'door_height': None,
        'door_width': None,
        'door_dia_diff': None
    }
    hole_sizes = {}
    if any('window' in key or 'door' in key for key in dimensions):
        background_color = (255, 255, 255)
        hole_sizes = extract_hole_line(new_cloud, background_color)
        ic(hole_sizes)
    else:
        print('no hole')

    if any('window' in key for key in dimensions):
        errors['window_width'] = hole_sizes.get('hole_bottom_len') - dimensions.get('window_width_bottom')
        errors['window_height'] = hole_sizes.get('hole_left_len') - dimensions.get('window_height_left')
        window_sizes['window_height'] = hole_sizes['hole_left_len']
        window_sizes['window_width'] = hole_sizes['hole_bottom_len']
        if dimensions['window_diagonal_bottom'] != 0 and  dimensions['window_diagonal_top'] != 0:
            cad_window_dia_diff = abs(dimensions.get('window_diagonal_bottom') - dimensions.get('window_diagonal_top'))
            errors['window_dia_diff'] = hole_sizes.get('hole_dia_diff') - cad_window_dia_diff
            window_sizes['window_dia_diff'] = hole_sizes['hole_dia_diff']

    if any('door' in key for key in dimensions):
        errors['door_width'] = hole_sizes.get('hole_bottom_len') - dimensions.get('door_width_bottom')
        errors['door_height'] = hole_sizes.get('hole_left_len') - dimensions.get('door_height_left')
        door_sizes['door_height'] = hole_sizes['hole_left_len']
        door_sizes['door_width'] = hole_sizes['hole_bottom_len']
        if dimensions['door_diagonal_bottom']!=0 and dimensions['door_diagonal_top']!=0:
            cad_door_dia_diff = abs(dimensions.get('door_diagonal_bottom') - dimensions.get('door_diagonal_top'))
            errors['door_dia_diff'] = hole_sizes.get('hole_dia_diff') - cad_door_dia_diff
            door_sizes['door_dia_diff'] = hole_sizes['hole_dia_diff']

    if dimensions['wall_diagonal_top'] != 0 and  dimensions['wall_diagonal_bottom'] != 0:
        cad_wall_dia_diff = abs(dimensions.get('wall_diagonal_top') - dimensions.get('wall_diagonal_bottom'))
        dia_diff1 = abs(abs(xs_argmax_x1 - xs_argmax_x2) - cad_wall_dia_diff)
        dia_diff2 = abs(wall_sizes.get('dia_diff') - cad_wall_dia_diff)
        errors['wall_dia_diff'] = min(dia_diff1, dia_diff2)

    ic(window_sizes)
    ic(door_sizes)
    ic(errors)



    # =================转json====================
    def Export2Json(dimensions, wall_sizes, hole_sizes, errors):
        with open(str(save_path / "results.json"), 'w') as file:
            json.dump({
                'dimensions': dimensions,
                'wall_final_sizes': wall_sizes,
                'window_sizes': window_sizes,
                'door_sizes': door_sizes,
                'hole_sizes': hole_sizes,
                'errors': errors
            }, file, indent=4)

    Export2Json(dimensions, wall_sizes, hole_sizes, errors)
    logger.success(f"json保存成功, 路径: {str(save_path)}")



    # =================数据写入表格====================
    # excel save path
    excel_save_path = str(save_path / "results.xlsx")
    # 写入平行线结果, 相差1度的斜率差约为0.018
    para_result = "合格" if k_diff <= 0.02 else "不合格"

    def editExcel(excel_file_path, wall_sizes, window_sizes, door_sizes,hole_sizes, errors, dimensions, para_result):
        try:
            workbook = openpyxl.load_workbook(excel_file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # 选择工作表
        sheet = workbook.active
        sheet.title = 'Sheet1'
        # 清空之前的内容
        for row in range(12, 20):
            for col in ['C', 'D', 'E', 'F']:
                cell = f'{col}{row}'
                for merged_range in sheet.merged_cells.ranges:
                    if cell in merged_range:
                        # 取消合并单元格
                        sheet.unmerge_cells(str(merged_range))
                        break
                sheet[cell] = ''  # 清空单元格内容

        # 获取当前日期
        current_date = datetime.now().strftime("%Y/%m/%d")  # 格式化日期为年/月/日形式

        # 写入标题
        sheet['C12'] = '墙扳高度'
        sheet['C13'] = '墙板高度'
        sheet['C14'] = '墙板宽度'
        sheet['C15'] = '墙板宽度'
        sheet['C16'] = '墙板对角线'
        sheet['F12'] = '(-3,0)'
        sheet['F13'] = '(-3,0)'
        sheet['F14'] = '(-3,0)'
        sheet['F15'] = '(-3,0)'
        sheet['F16'] = '△≤3'
        if any('window' in key for key in dimensions):
            sheet['C17'] = '窗洞高度'
            sheet['C18'] = '窗洞宽度'
            sheet['C19'] = '窗洞对角线'
            sheet['F17'] = '(0,5)'
            sheet['F18'] = '(0,5)'
            sheet['F19'] = '△≤5.0'
        if any('door' in key for key in dimensions):
            sheet['C17'] = '门洞高度'
            sheet['C18'] = '门洞宽度'
            sheet['C19'] = '门洞对角线'
            sheet['F17'] = '(0,5)'
            sheet['F18'] = '(0,5)'
            sheet['F19'] = '△≤5.0'
        sheet['C20'] = '外墙平行度'

        # 定义单元格映射
        cell_mapping = {
            'time': 'F4',
            'wall_height_left': 'D12',
            'wall_height_right': 'D13',
            'wall_width_bottom': 'D14',
            'wall_width_top': 'D15',
            'wall_diagonal_diff': 'D16'
        }

        if any('window' in key for key in dimensions):
            cell_mapping2 = {
                'window_width_bottom': 'D17',
                'window_height_left': 'D18',
                'window_diagonal_diff': 'D19'
            }

        if any('door' in key for key in dimensions):
            cell_mapping2 = {
                'door_width_bottom': 'D17',
                'door_height_left': 'D18',
                'door_diagonal_diff': 'D19'
            }

        # 定义其他单元格映射
        cell_mapping3 = {
            'wall_left_height': 'E12',
            'wall_right_height': 'E13',
            'wall_bottom_width': 'E14',
            'wall_top_width': 'E15',
            'wall_dia_diff': 'E16',
            'window_width': 'E17',
            'window_height': 'E18',
            'window_dia_diff': 'E19',
            'door_width': 'E17',
            'door_height': 'E18',
            'door_dia_diff': 'E19'
        }

        # 写入CAD值
        for key, cell in cell_mapping.items():
            value = dimensions.get(key)
            if value is not None:
                if isinstance(value, (int, float)):
                    sheet[cell] = '{:.2f}'.format(value)
                else:
                    sheet[cell] = value
            elif key == 'time':
                sheet[cell] = current_date
        if 'cell_mapping2' in locals() or 'cell_mapping2' in globals():
            for key, cell in cell_mapping2.items():
                value = dimensions.get(key)
                if value is not None:
                    if isinstance(value, (int, float)):
                        sheet[cell] = '{:.2f}'.format(value)
                    else:
                        sheet[cell] = value
        # 将检查结果写入
        for key, cell in cell_mapping3.items():
            value = errors.get(key)
            if value is not None:
                if isinstance(value, (int, float)):
                    # 根据条件判断合格性
                    if key in ['wall_left_height', 'wall_right_height', 'wall_bottom_width', 'wall_top_width']:
                        if -6 < value < 3:
                            sheet[cell] = '合格'
                        else:
                            sheet[cell] = '不合格'
                    elif key in ['wall_dia_diff']:
                        if 0 < value < 6:
                            sheet[cell] = '合格'
                        else:
                            sheet[cell] = '不合格'
                    elif key in ['window_width', 'window_height', 'window_dia_diff']:
                        if -3 < value < 8:
                            sheet[cell] = '合格'
                        else:
                            sheet[cell] = '不合格'
                    elif key in ['door_width', 'door_height', 'door_dia_diff']:
                        if -3 < value < 8:
                            sheet[cell] = '合格'
                        else:
                            sheet[cell] = '不合格'
                else:
                    sheet[cell] = value

        from openpyxl.styles import Font, Alignment, Border, Side

        # 定义边框样式
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        sheet['D20'] = para_result
        # 检查D，E列是否存在空白单元格，删除空白单元格整行
        for row in range(19, 11, -1):
            d_value = sheet[f'D{row}'].value
            e_value = sheet[f'E{row}'].value
            print(f"Checking row {row}: D{row} = {d_value}, E{row} = {e_value}")
            if d_value is None or (isinstance(d_value, str) and d_value.strip() == '') or \
                    e_value is None or (isinstance(e_value, str) and e_value.strip() == ''):
                print(f"Deleting row {row} because D{row} or E{row} is empty.")
                sheet.delete_rows(row)

        #更新 last_row
        last_row = sheet.max_row
        sheet.merge_cells(start_row=last_row, start_column=4, end_row=last_row, end_column=6)
        sheet[f'D{last_row}'].alignment = Alignment(horizontal="center", vertical="center")

        #填写序号
        for row in range(12, sheet.max_row + 1):
            sheet[f'B{row}'] = row - 11

        # 设置单元格格式
        for row in sheet.iter_rows(min_row=12, max_row=20, min_col=2, max_col=6):
            for cell in row:
                if cell.value is not None:
                    cell.font = Font(name="宋体", size=6.5, bold=True)
                    cell.alignment = Alignment(vertical="center", horizontal="center", wrap_text=True)
                    cell.border = border_style

        # 保存修改后的 Excel 文件
        workbook.save(excel_file_path)
        print("空白单元格已清空，格式已统一，边框已添加。")

    editExcel(EXCEL_TEMPLATE, wall_sizes, window_sizes, door_sizes,hole_sizes, errors, dimensions, para_result)
    logger.success(f"数据写入excel成功, 路径: {excel_save_path}")



    # =================cad对齐====================
    new_cloud.scale(1000, 1000, 1000)
    img_grey_bg_cad, img_white_bg_cad = cad_align(new_cloud, dimensions, cad_model_path, cad_img=True)
    cv2.imwrite(str(save_path / "img_grey_bg_cad.png"), img_grey_bg_cad)
    cv2.imwrite(str(save_path / "img_white_bg_cad.png"), img_white_bg_cad)

    img_grey_bg, img_white_bg = cad_align(new_cloud, dimensions, cad_model_path, cad_img=False)
    img_white_bg, img_grey_bg = draw_text_on_img(img_white_bg, img_grey_bg, wall_sizes, window_sizes, door_sizes)

    # 保存图片
    cv2.imwrite(str(save_path / "img_grey_bg.png"), img_grey_bg)
    cv2.imwrite(str(save_path / "img_white_bg.png"), img_white_bg)
    logger.success(f"cad对齐图片保存成功, 路径: {str(save_path)}")



    # =================图片插入excel====================
    imgPath = str(save_path / "img_white_bg.png")
    # insertimg2excel(imgPath, excelPath)

    wb = load_workbook(EXCEL_TEMPLATE)
    ws = wb.active
    img = Image(imgPath)
    ratio = img.width/img.height
    img.height = (ws.row_dimensions[10].height)*1.2
    img.width = ratio * img.height
    ws.add_image(img, 'E10')
    ws['E10'].alignment = Alignment(horizontal="center", vertical="center")
    wb.save(excel_save_path)
    logger.success(f"图片插入excel成功, 路径: {excel_save_path}")



    # =================转pdf====================
    import win32com.client
    import os
    excel = win32com.client.Dispatch("Excel.Application")
    excel.Visible = False

    current_dir = os.getcwd()
    workbook = excel.Workbooks.Open(os.path.join(current_dir, excel_save_path))
    workbook.ExportAsFixedFormat(0, os.path.join(current_dir, str(save_path / "results.pdf")))

    workbook.Close()
    excel.Quit()
    logger.success(f"pdf保存成功, 路径: {str(save_path)}")


if __name__ == "__main__":
    # single measurement
    cloud_path = r"D:\0309_test_wall\11am_wall\combined_pcd.ply"
    cad_model_path = r"D:\0308_test_wall\J4_2025-2-19_LINE.dxf"
    all_measurement(cloud_path, cad_model_path)

    # measure all in one folder
    # pointcloud_dir = r"Z:\data\Project_Hebei\20250103\pointcloud"
    # cad_dir = r"Z:\data\Project_Hebei\20250103\cad"
    # for pointcloud_path in list(Path(pointcloud_dir).glob("*.ply")):
    #     cad_path = Path(cad_dir) / (pointcloud_path.stem + ".dxf")
    #     logger.info(f"开始处理点云: {pointcloud_path}, CAD模型: {cad_path}")
    #     all_measurement(str(pointcloud_path), str(cad_path))
    #     logger.success(f"处理完成, 路径: {pointcloud_path}")
    #     print("\n"*5)
